from openpyxl.workbook import Workbook
from openpyxl import load_workbook

#Loading in existing spreadsheet
wb = load_workbook('job_roles_list.xlsx')

#Creating an active worksheet
ws = wb.active

#Print an employee job role from our Job_Roles
print(f' Employee #{ws['A2'].value} is a {ws['B2'].value}')

print(f'---------------------------------------------')

#Grab a whole column
#column_a = ws['A']
#print(column_a)

print(f'---------------------------------------------')
 
#Grabbing the first 10 employees id#, job role and job department
range = ws['A2':'C11']
for x in range:
    for y in x:
        print(y.value)

    